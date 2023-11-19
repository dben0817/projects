import requests
from openpyxl import Workbook

def search_books(subject, filter_free=False):
    # Send an HTTP GET request to the Google Books API
    url = 'https://www.googleapis.com/books/v1/volumes'
    params = {
        'q': f'subject:{subject}',
        'maxResults': 40,  # Increase the maxResults to fetch more books (max 40 per API request)
        'key': 'See comment ->' # TODO: This is where you insert your API key for the Google Books services
    }

    if filter_free:
        params['filter'] = 'free-ebooks'

    response = requests.get(url, params=params)
    data = response.json()

    # Process the response data
    if 'items' in data:
        # Create a new Excel workbook and get the active sheet
        workbook = Workbook()
        sheet = workbook.active

        # Set column headers
        sheet['A1'] = 'Title'
        sheet['B1'] = 'Authors'
        sheet['C1'] = 'Publisher'
        sheet['D1'] = 'Published Date'
        sheet['E1'] = 'Page Count'
        sheet['F1'] = 'ISBN'

        row = 2  # Start from the second row

        for book in data['items']:
            volume_info = book.get('volumeInfo', {})
            title = volume_info.get('title', '')
            authors = ', '.join(volume_info.get('authors', ['Unknown']))
            publisher = volume_info.get('publisher', '')
            published_date = volume_info.get('publishedDate', '')
            page_count = volume_info.get('pageCount', '')
            isbn = volume_info.get('industryIdentifiers', [{}])[0].get('identifier', '')

            # Write book details to the spreadsheet
            sheet.cell(row=row, column=1, value=title)
            sheet.cell(row=row, column=2, value=authors)
            sheet.cell(row=row, column=3, value=publisher)
            sheet.cell(row=row, column=4, value=published_date)
            sheet.cell(row=row, column=5, value=page_count)
            sheet.cell(row=row, column=6, value=isbn)
            row += 1

        # Save the workbook to a file
        workbook.save(f'{subject}_books.xlsx')
        print(f'{len(data["items"])} books saved to {subject}_books.xlsx')
    else:
        print(f'No {subject} books found.')

# Prompt the user for input
subject = input('Enter the subject you want to search for: ')
filter_free = input('Filter for free books? (y/n): ').lower() == 'y'

# Perform the book search and save to a spreadsheet
search_books(subject, filter_free)
